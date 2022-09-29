from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

headings = ('Income/Month', 'Fixed Expenses', 'Purchase History')
purchase_history = ('Date', 'Expenses', 'Amount', 'Budget Balance')
def create_profile(name):
    budget = Workbook()
    budge = budget.active
    budge.title = name
    budge.merge_cells('A1:B1')
    budge.merge_cells('D1:E1')
    budge.merge_cells('G1:J1')
    char = get_column_letter
    budge[char(1) + '1'].fill = PatternFill(fill_type='solid', start_color='0000FF00', end_color='0000FF00')
    budge[char(4) + '1'].fill = PatternFill(fill_type='solid', start_color='00FF9900', end_color='00FF9900')
    budge[char(7) + '1'].fill = PatternFill(fill_type='solid', start_color='0000CCFF', end_color='0000CCFF')
    for i in headings:
        if headings.index(i) == 0: a = headings.index(i) + 1
        else: a = a + 3
        budge[char(a) + '1'] = i
    budge[char(7) + '1'].alignment = Alignment(horizontal='center')
    for i in purchase_history:
        budge[char(7 + purchase_history.index(i)) + '2'] = i
    budge.column_dimensions[char(10)].width = len('Budget Balance')
    budget.save(f'{name}.xlsx')
class Category:
    def __init__(self,field, file):
        self.field = field
        self.budget = load_workbook(f'{file}.xlsx')
        self.budge = self.budget.active
        self.storage = {}
        self.timestamps = []
        self.check_cells(self.field)
    # For positioning data to their respective fields
    def field_position(self, field, rows, col=False):
        char = get_column_letter
        if field == 'income':
            if col == True:
                return char(2) + str(rows+2)
            return char(1) + str(rows+2)
        elif field == 'expenses':
            if col == True:
                return char(5) + str(rows+2)
            return char(4) + str(rows+2)
        else:
            if col == True:
                return char(9) + str(rows+3)
            return char(8) + str(rows+3)

    # For checking cells for existing data
    def check_cells(self,field):
        char = self.field_position
        for row in range(1000):
            if self.budge[char(self.field, row)].value == None or self.budge[char(self.field, row)].value == 'Total':
                break
            self.storage[self.budge[char(self.field, row)].value] = self.budge[char(self.field, row, True)].value
            # For checking timestamps
            if self.field == 'spending':
                if self.budge['G' + str(row+3)] == '':
                    continue
                else:
                    self.timestamps.append(self.budge['G' + str(row+3)].value)
    def add_entry(self, description, money):
        char = self.field_position
        self.storage[f'{description}'] = int(money)
        keys = list(self.storage.keys())
        if self.field == 'spending':
            self.timestamps.append(datetime.today().strftime('%Y-%m-%d'))
        for rows in range(len(self.storage)):
            self.budge[char(self.field, rows, True)] = self.storage[keys[rows]]
            self.budge[char(self.field, rows)] = keys[rows]
            if self.field == 'spending':
                self.budge['G' + str(rows+3)] = self.timestamps[rows]
        self.storage['Total'] = 0
        total = (t for t in self.storage.values())
        self.storage['Total'] = sum(total)
        keys.append('Total')
        self.budge[char(self.field, len(keys)-1, True)] = self.storage['Total']
        self.budge[char(self.field, len(keys)-1)] = keys[len(keys)-1]
        self.budget.save(f'{file}.xlsx')

choice = input('Do you already have a profile? yes or no: ')
if choice == 'yes':
    file = input('What is the file name?: ')
elif choice == 'no':
    file = input('create your profile: ')
    create_profile(file)
while True:
    print("Do you want to add source of income(1), add fixed expenses(2), add today's purchases(3)?, or exit(4)")
    choice = input("input choice number to do the following: ")
    if choice == '1':
        description = input('What is your source of income?: ')
        money = input('How much do you earn in a month?:')
        choice = Category('income', file)
        choice.add_entry(description, money)
    elif choice == '2':
        description = input('What are your monthly expenses?: ')
        money = input('Monthly expense amount?:')
        choice = Category('expenses', file)
        choice.add_entry(description, money)
    elif choice == '3':
        description = input('What did you spend today?: ')
        money = input('How much was it?:')
        choice = Category('spending', file)
        choice.add_entry(description, money)
        print(choice.timestamps)
    else:
        exit()